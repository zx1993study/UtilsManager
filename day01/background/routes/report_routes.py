
"""
报告管理路由
"""
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import get_db
from models.api_result import ApiResult
from models.page_result import PageResult
from models.flow_result import FlowResult
from logger_config import logger
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from tempfile import NamedTemporaryFile

report_bp = Blueprint('reports', __name__)

@report_bp.route('/api/reports', methods=['GET'])
@jwt_required()
def get_reports():
    """获取报告列表"""
    try:
        skip = int(request.args.get('skip', 0))
        limit = int(request.args.get('limit', 100))
        report_type = request.args.get('type')
        status = request.args.get('status')
        date_from = request.args.get('dateFrom')
        date_to = request.args.get('dateTo')

        db = next(get_db())

        # 根据类型获取不同的报告
        if report_type == 'interface':
            query = db.query(ApiResult)
        elif report_type == 'page':
            query = db.query(PageResult)
        elif report_type == 'flow':
            query = db.query(FlowResult)
        else:
            # 获取所有类型的报告
            api_results = [r.to_dict() for r in db.query(ApiResult).offset(skip).limit(limit).all()]
            page_results = [r.to_dict() for r in db.query(PageResult).offset(skip).limit(limit).all()]
            flow_results = [r.to_dict() for r in db.query(FlowResult).offset(skip).limit(limit).all()]

            all_results = api_results + page_results + flow_results
            all_results.sort(key=lambda x: x['createTime'], reverse=True)

            return jsonify(all_results[:limit]), 200

        # 应用过滤条件
        if status:
            query = query.filter(ApiResult.result_status == status if report_type == 'interface' else
                              PageResult.result_status == status if report_type == 'page' else
                              FlowResult.status == status)

        if date_from:
            query = query.filter(ApiResult.create_time >= datetime.fromisoformat(date_from) if report_type == 'interface' else
                               PageResult.create_time >= datetime.fromisoformat(date_from) if report_type == 'page' else
                               FlowResult.create_time >= datetime.fromisoformat(date_from))

        if date_to:
            query = query.filter(ApiResult.create_time <= datetime.fromisoformat(date_to) if report_type == 'interface' else
                               PageResult.create_time <= datetime.fromisoformat(date_to) if report_type == 'page' else
                               FlowResult.create_time <= datetime.fromisoformat(date_to))

        results = query.offset(skip).limit(limit).all()

        return jsonify([result.to_dict() for result in results]), 200

    except Exception as e:
        logger.error(f'获取报告列表失败: {str(e)}')
        return jsonify({'message': '获取报告列表失败'}), 500

@report_bp.route('/api/reports/<report_type>/<int:result_id>', methods=['GET'])
@jwt_required()
def get_report(report_type, result_id):
    """获取单个报告详情"""
    try:
        db = next(get_db())

        if report_type == 'interface':
            result = db.query(ApiResult).filter(ApiResult.api_result_id == result_id).first()
        elif report_type == 'page':
            result = db.query(PageResult).filter(PageResult.page_result_id == result_id).first()
        elif report_type == 'flow':
            result = db.query(FlowResult).filter(FlowResult.flow_result_id == result_id).first()
        else:
            return jsonify({'message': '无效的报告类型'}), 400

        if not result:
            return jsonify({'message': '报告不存在'}), 404

        return jsonify(result.to_dict()), 200

    except Exception as e:
        logger.error(f'获取报告详情失败: {str(e)}')
        return jsonify({'message': '获取报告详情失败'}), 500

@report_bp.route('/api/reports/export', methods=['POST'])
@jwt_required()
def export_reports():
    """导出报告为Excel"""
    try:
        data = request.get_json()
        report_type = data.get('type', 'all')
        date_from = data.get('dateFrom')
        date_to = data.get('dateTo')

        db = next(get_db())

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "测试报告"

        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # 根据类型设置表头
        if report_type == 'interface':
            headers = ['ID', '实例ID', '返回信息', '状态', 'Code', '备注', '创建时间']
            query = db.query(ApiResult)
        elif report_type == 'page':
            headers = ['ID', '页面实例ID', '返回信息', '状态', 'Code', '截图路径', '备注', '创建时间']
            query = db.query(PageResult)
        elif report_type == 'flow':
            headers = ['ID', '流程ID', '结果ID', '预期结果', '状态', '备注', '创建时间']
            query = db.query(FlowResult)
        else:
            return jsonify({'message': '无效的报告类型'}), 400

        # 应用日期过滤
        if date_from:
            query = query.filter(ApiResult.create_time >= datetime.fromisoformat(date_from) if report_type == 'interface' else
                               PageResult.create_time >= datetime.fromisoformat(date_from) if report_type == 'page' else
                               FlowResult.create_time >= datetime.fromisoformat(date_from))

        if date_to:
            query = query.filter(ApiResult.create_time <= datetime.fromisoformat(date_to) if report_type == 'interface' else
                               PageResult.create_time <= datetime.fromisoformat(date_to) if report_type == 'page' else
                               FlowResult.create_time <= datetime.fromisoformat(date_to))

        results = query.all()

        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # 写入数据
        for row_num, result in enumerate(results, 2):
            result_dict = result.to_dict()
            for col_num, key in enumerate(['id', 'instanceId' if report_type != 'flow' else 'flowId', 
                                           'returnInfo' if report_type != 'flow' else 'resultId',
                                           'resultStatus' if report_type != 'flow' else 'expectResult',
                                           'code' if report_type != 'flow' else 'status',
                                           'screenshotPath' if report_type == 'page' else 'remark',
                                           'remark' if report_type != 'page' else 'createTime'], 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = result_dict.get(key, '')
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column[0].column_letter].width = min(adjusted_width, 50)

        # 保存到临时文件
        with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name

        logger.info(f'导出报告成功: {report_type}')

        return send_file(
            tmp_path,
            as_attachment=True,
            download_name=f'test_report_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f'导出报告失败: {str(e)}')
        return jsonify({'message': '导出报告失败'}), 500
